"""Parse uploaded files into plain text for LLM context."""
import csv
import io
import logging
import os

logger = logging.getLogger(__name__)


def parse_file(storage_path: str, file_type: str) -> str:
    """Extract text content from a file. Returns empty string on failure."""
    try:
        parser = _PARSERS.get(file_type)
        if not parser:
            logger.warning(f"No parser for file type: {file_type}")
            return ""
        return parser(storage_path)
    except Exception as e:
        logger.error(f"Failed to parse file {storage_path}: {e}")
        return ""


def _parse_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _parse_pdf(path: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(path)
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def _parse_docx(path: str) -> str:
    from docx import Document

    doc = Document(path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_csv(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = []
        for row in reader:
            rows.append(" | ".join(row))
        return "\n".join(rows)


def _parse_xlsx(path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"[Sheet: {ws.title}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets)


_PARSERS = {
    "txt": _parse_txt,
    "md": _parse_txt,
    "pdf": _parse_pdf,
    "docx": _parse_docx,
    "csv": _parse_csv,
    "xlsx": _parse_xlsx,
}
